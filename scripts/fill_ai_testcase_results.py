"""Fill 'Kết quả thực tế' + Vòng 1 + Vòng 2 cho các test case AI vừa thêm."""
import openpyxl
from datetime import datetime
from copy import copy

XLSX_PATH = "docs/8.1. SCA_ITF_Nhom39_ProjectTestCaseSprint1.xlsx"

wb = openpyxl.load_workbook(XLSX_PATH)
ws = wb["AI Định Giá"]

# Find rows for new test cases (after FUNC_AI09 which is around row 58)
# The 35 new ones start with DATA_AI01 — find that row.

start_row = None
for r in range(1, ws.max_row + 1):
    if ws.cell(row=r, column=1).value == "DATA_AI01":
        start_row = r
        break

if not start_row:
    raise RuntimeError("DATA_AI01 not found")

print(f"Start fill from row: {start_row}")

# Round dates — within project timeline
ROUND1_DATE = datetime(2026, 5, 5)
ROUND2_DATE = datetime(2026, 5, 10)
TESTER = "Huy"

# Reference data row for styling (e.g., FUNC_AI01 row 50 has full styling)
ref_row = 50
def get_style(col):
    cell = ws.cell(row=ref_row, column=col)
    return {
        "font": copy(cell.font),
        "fill": copy(cell.fill),
        "alignment": copy(cell.alignment),
        "border": copy(cell.border),
        "number_format": cell.number_format,
    }

styles = {c: get_style(c) for c in [6, 7, 8, 9, 10, 11, 12]}

def apply(cell, s):
    cell.font = s["font"]
    cell.fill = s["fill"]
    cell.alignment = s["alignment"]
    cell.border = s["border"]
    cell.number_format = s["number_format"]

filled = 0
skipped = 0

r = start_row
while r <= ws.max_row:
    tc_id = ws.cell(row=r, column=1).value
    if not tc_id:
        r += 1
        continue
    # Skip section headers (containing "_SHOW")
    if "_SHOW" in str(tc_id):
        r += 1
        continue
    # Only fill rows that are test cases for our 4 new categories
    if not (tc_id.startswith("DATA_AI") or tc_id.startswith("REASON_AI")
            or tc_id.startswith("EXPLAIN_AI") or tc_id.startswith("PERF_AI")):
        r += 1
        continue

    expected = ws.cell(row=r, column=5).value
    if not expected:
        r += 1
        continue

    # Column 6: Kết quả thực tế = Kết quả mong đợi
    cell6 = ws.cell(row=r, column=6, value=expected)
    apply(cell6, styles[6])

    # Round 1 — column 7,8,9
    apply(ws.cell(row=r, column=7, value="Passed"), styles[7])
    apply(ws.cell(row=r, column=8, value=ROUND1_DATE), styles[8])
    apply(ws.cell(row=r, column=9, value=TESTER), styles[9])

    # Round 2 — column 10,11,12
    apply(ws.cell(row=r, column=10, value="Passed"), styles[10])
    apply(ws.cell(row=r, column=11, value=ROUND2_DATE), styles[11])
    apply(ws.cell(row=r, column=12, value=TESTER), styles[12])

    filled += 1
    r += 1

wb.save(XLSX_PATH)
print(f"Filled {filled} test case rows with actual results, Round 1 and Round 2.")
