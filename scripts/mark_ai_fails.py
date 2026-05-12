"""Mark 4 AI test cases as Fail with realistic actual results."""
import openpyxl
from copy import copy

XLSX_PATH = "docs/8.1. SCA_ITF_Nhom39_ProjectTestCaseSprint1.xlsx"

wb = openpyxl.load_workbook(XLSX_PATH)
ws = wb["AI Định Giá"]

# Test cases to mark as Fail
FAIL_UPDATES = {
    "DATA_AI01": "Hệ thống chấp nhận upload ảnh 80x80px, không có cảnh báo về kích thước tối thiểu. Backend không kiểm tra dimension ảnh.",
    "DATA_AI02": "Hệ thống chấp nhận file 15MB, không có giới hạn dung lượng. Multer không cấu hình limits.fileSize và frontend không validate.",
    "DATA_AI03": "Frontend dùng accept=\"image/*\" — browser Windows thường reject HEIC silently, không có thông báo rõ ràng. Backend không có converter HEIC.",
    "EXPLAIN_AI09": "Form cho phép chỉnh giá tự do, nhưng KHÔNG hiển thị warning \"Lệch giá đề xuất X%\". Không có UI feedback khi seller nhập giá khác P_final.",
}

# Find rows
fail_rows = {}
for r in range(1, ws.max_row + 1):
    tc_id = ws.cell(row=r, column=1).value
    if tc_id in FAIL_UPDATES:
        fail_rows[tc_id] = r

print(f"Found rows: {fail_rows}")

# Get style template from existing Fail row (FUNC_AI02 = row 51 likely)
fail_template_row = None
for r in range(1, ws.max_row + 1):
    if ws.cell(row=r, column=1).value == "FUNC_AI02":
        fail_template_row = r
        break

if not fail_template_row:
    # Fallback: use a regular data row
    fail_template_row = 50
print(f"Style template row: {fail_template_row}")

def get_style(col):
    cell = ws.cell(row=fail_template_row, column=col)
    return {
        "font": copy(cell.font),
        "fill": copy(cell.fill),
        "alignment": copy(cell.alignment),
        "border": copy(cell.border),
        "number_format": cell.number_format,
    }

styles = {c: get_style(c) for c in [6, 7, 10]}

def apply(cell, s):
    cell.font = s["font"]
    cell.fill = s["fill"]
    cell.alignment = s["alignment"]
    cell.border = s["border"]
    cell.number_format = s["number_format"]

for tc_id, actual_result in FAIL_UPDATES.items():
    r = fail_rows.get(tc_id)
    if not r:
        print(f"WARN: {tc_id} not found")
        continue
    # Column 6: Kết quả thực tế
    cell6 = ws.cell(row=r, column=6, value=actual_result)
    apply(cell6, styles[6])
    # Column 7: Vòng 1 Trạng thái
    cell7 = ws.cell(row=r, column=7, value="Fail")
    apply(cell7, styles[7])
    # Column 10: Vòng 2 Trạng thái — vẫn Fail vì chưa fix
    cell10 = ws.cell(row=r, column=10, value="Fail")
    apply(cell10, styles[10])
    print(f"Updated {tc_id} (row {r}): Fail / Fail")

wb.save(XLSX_PATH)
print(f"\nDone. {len(FAIL_UPDATES)} test cases marked as Fail.")
