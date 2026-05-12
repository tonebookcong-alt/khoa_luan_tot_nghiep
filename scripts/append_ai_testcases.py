"""Append additional AI pricing test cases to the testcase xlsx."""
import openpyxl
from copy import copy

XLSX_PATH = "docs/8.1. SCA_ITF_Nhom39_ProjectTestCaseSprint1.xlsx"

wb = openpyxl.load_workbook(XLSX_PATH)
ws = wb["AI Định Giá"]

last_row = ws.max_row
print(f"Last row before append: {last_row}")

header_template_row = 49  # FUNCTION_SHOW row
data_template_row = 50    # FUNC_AI01 row

def get_row_styles(row_idx, ncols=13):
    styles = []
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row_idx, column=c)
        styles.append({
            "font": copy(cell.font),
            "fill": copy(cell.fill),
            "alignment": copy(cell.alignment),
            "border": copy(cell.border),
        })
    return styles

header_styles = get_row_styles(header_template_row)
data_styles = get_row_styles(data_template_row)

def apply_style(cell, style):
    cell.font = style["font"]
    cell.fill = style["fill"]
    cell.alignment = style["alignment"]
    cell.border = style["border"]

def write_section_header(row_idx, label):
    ws.cell(row=row_idx, column=1, value=label)
    for c in range(1, 14):
        apply_style(ws.cell(row=row_idx, column=c), header_styles[c - 1])

def write_test_case(row_idx, tc_id, desc, action, precond, expected):
    values = [tc_id, desc, action, precond, expected,
              None, None, None, None, None, None, None, None]
    for c, v in enumerate(values, start=1):
        cell = ws.cell(row=row_idx, column=c, value=v)
        apply_style(cell, data_styles[c - 1])

categories = [
    {
        "header": "DATA_SHOW Kiểm thử dữ liệu đầu vào",
        "cases": [
            ("DATA_AI01", "Ảnh kích thước rất nhỏ (< 100x100px)",
             "1. Tải lên ảnh kích thước 80x80px.\n2. Nhập đủ thông tin.\n3. Bấm \"Định giá bằng AI\".",
             "Truy cập trang Định giá AI",
             "Hệ thống cảnh báo ảnh quá nhỏ, độ tin cậy thấp HOẶC từ chối với thông báo yêu cầu ảnh tối thiểu."),
            ("DATA_AI02", "Ảnh dung lượng vượt giới hạn (> 10MB)",
             "1. Tải lên ảnh dung lượng 15MB.",
             "Truy cập trang Định giá AI",
             "Hệ thống chặn upload, thông báo \"Kích thước ảnh vượt giới hạn cho phép\"."),
            ("DATA_AI03", "Ảnh có định dạng HEIC (iOS)",
             "1. Tải lên ảnh .heic chụp từ iPhone.\n2. Nhập đủ thông tin.\n3. Bấm \"Định giá\".",
             "Truy cập trang Định giá AI",
             "Hệ thống chấp nhận và xử lý đúng, HOẶC từ chối với thông báo rõ ràng về định dạng hỗ trợ."),
            ("DATA_AI04", "Ảnh không phải điện thoại (xe, mèo, người…)",
             "1. Tải lên 5 ảnh không liên quan tới điện thoại.\n2. Nhập đủ thông tin.\n3. Bấm \"Định giá\".",
             "Truy cập trang Định giá AI",
             "YOLO không nhận diện được điện thoại, hệ thống hiển thị cảnh báo \"Không tìm thấy thiết bị trong ảnh\" và không tính P_final."),
            ("DATA_AI05", "Pin = 0%",
             "1. Tải đủ ảnh.\n2. Nhập % Pin = 0.\n3. Bấm \"Định giá\".",
             "Truy cập trang Định giá AI",
             "Báo lỗi \"Giá trị Pin phải trong khoảng 1-100\", không gửi request."),
            ("DATA_AI06", "Pin = 100% (biên trên hợp lệ)",
             "1. Tải đủ ảnh.\n2. Nhập % Pin = 100.\n3. Bấm \"Định giá\".",
             "Truy cập trang Định giá AI",
             "Yêu cầu được gửi đi, trả về kết quả định giá. Không khấu trừ do pin."),
            ("DATA_AI07", "Pin chứa ký tự đặc biệt (abc, 80%, -50)",
             "1. Tải đủ ảnh.\n2. Nhập % Pin = \"abc\" hoặc \"-50\".\n3. Bấm \"Định giá\".",
             "Truy cập trang Định giá AI",
             "Hệ thống chặn input không hợp lệ, hiển thị thông báo lỗi định dạng."),
            ("DATA_AI08", "Model khai báo không khớp với ảnh YOLO nhận diện",
             "1. Tải ảnh iPhone 14.\n2. Chọn Model = \"iPhone 6\".\n3. Bấm \"Định giá\".",
             "Truy cập trang Định giá AI",
             "Hệ thống hiển thị banner cảnh báo \"Model bạn khai báo có thể không khớp với máy trong ảnh\"."),
            ("DATA_AI09", "Upload ảnh trùng lặp (cùng 1 file 5 lần)",
             "1. Tải lên 5 lần cùng 1 ảnh.\n2. Nhập đủ thông tin.\n3. Bấm \"Định giá\".",
             "Truy cập trang Định giá AI",
             "Hệ thống vẫn chạy nhưng kết quả tính dựa trên 5 ảnh trùng (hoặc cảnh báo trùng)."),
            ("DATA_AI10", "Ảnh chụp ngược sáng / mờ / quá tối",
             "1. Tải lên 5 ảnh chất lượng kém (mờ, ngược sáng).\n2. Bấm \"Định giá\".",
             "Truy cập trang Định giá AI",
             "Confidence score thấp, có thể có banner gợi ý chụp lại ảnh."),
        ]
    },
    {
        "header": "REASON_SHOW Kiểm thử độ hợp lý của kết quả định giá",
        "cases": [
            ("REASON_AI01", "Giá đề xuất luôn dương và không vượt giá thị trường",
             "1. Định giá iPhone 14 cũ.\n2. Quan sát P_final so với P_market.",
             "Định giá thành công",
             "P_final > 0 và P_final ≤ P_market (do khấu trừ % hư hỏng)."),
            ("REASON_AI02", "Máy hư hỏng nặng có giá thấp hơn máy mới",
             "1. Định giá 2 lần cùng model: lần 1 ảnh máy mới, lần 2 ảnh máy nhiều trầy/nứt.",
             "Định giá thành công",
             "P_final lần 2 < P_final lần 1."),
            ("REASON_AI03", "Tổng damage weights bằng 1.0",
             "1. Xem JSON kết quả damageBreakdown.\n2. Cộng tất cả weight (screen 0.40 + battery 0.20 + physical 0.25 + camera 0.15 + other 0.05).",
             "Định giá thành công",
             "Tổng weight = 1.0 (kiểm chứng công thức trọng số đầy đủ)."),
            ("REASON_AI04", "Confidence score nằm trong [0, 1]",
             "1. Định giá nhiều lần với data khác nhau.\n2. Quan sát confidence score.",
             "Định giá thành công",
             "0 ≤ confidenceScore ≤ 1 cho mọi trường hợp."),
            ("REASON_AI05", "Khoảng giá [low, high] bao quanh P_final",
             "1. Định giá thành công, đọc priceRange.",
             "Định giá thành công",
             "priceRange.low ≤ P_final ≤ priceRange.high. Tỉ lệ ±8% so với P_final."),
            ("REASON_AI06", "iPhone X trở lên có giá; iPhone 6/7/8 chỉ \"tham khảo\"",
             "1. Định giá iPhone 6.\n2. Định giá iPhone X.",
             "Định giá thành công",
             "iPhone 6: hiển thị \"Giá tham khảo < 2tr\" không có P_final cụ thể.\niPhone X: trả về P_final đầy đủ."),
            ("REASON_AI07", "Hai ảnh giống nhau cho kết quả tương đương",
             "1. Định giá cùng 1 bộ ảnh 2 lần liên tiếp.",
             "Định giá thành công",
             "Hai lần định giá có P_final chênh lệch nhỏ (< 5%) — kiểm tra ổn định mô hình."),
            ("REASON_AI08", "P_market là median, không bị giá rác kéo lên/xuống",
             "1. Định giá iPhone phổ biến.\n2. Kiểm tra ai_analysis_log MongoDB.",
             "Có dữ liệu market_price_raw trong MongoDB",
             "P_market = median của dataset, không phải mean. Loại bỏ outlier giá quá thấp/cao."),
        ]
    },
    {
        "header": "EXPLAIN_SHOW Kiểm thử giải thích kết quả",
        "cases": [
            ("EXPLAIN_AI01", "Hiển thị giá thị trường (P_market)",
             "1. Định giá thành công.\n2. Quan sát màn hình kết quả.",
             "Định giá thành công",
             "Trang kết quả hiển thị rõ \"Giá thị trường tham khảo: X VND\" cùng nguồn dữ liệu (Chợ Tốt)."),
            ("EXPLAIN_AI02", "Hiển thị giá đề xuất cuối (P_final)",
             "1. Định giá thành công.",
             "Định giá thành công",
             "Hiển thị nổi bật \"Giá đề xuất: Y VND\" với màu sắc phân biệt."),
            ("EXPLAIN_AI03", "Hiển thị khoảng giá hợp lý [low, high]",
             "1. Định giá thành công.",
             "Định giá thành công",
             "Hiển thị \"Khoảng giá hợp lý: X - Y VND\" có dải tin cậy ±8%."),
            ("EXPLAIN_AI04", "Phân tích chi tiết hư hỏng theo từng bộ phận",
             "1. Định giá ảnh có hư hỏng.\n2. Mở phần breakdown.",
             "Định giá thành công",
             "Hiển thị bảng: Màn hình / Pin / Vỏ máy / Camera / Khác, mỗi mục có % khấu trừ, mô tả."),
            ("EXPLAIN_AI05", "Hiển thị bounding box trên ảnh",
             "1. Định giá ảnh có damage.\n2. Xem ảnh trong kết quả.",
             "Định giá thành công",
             "Ảnh hiển thị bounding box quanh vùng hư hỏng kèm nhãn (scratch / physical_damage / screen_defect)."),
            ("EXPLAIN_AI06", "Hiển thị độ tin cậy (confidence score)",
             "1. Định giá nhiều case.",
             "Định giá thành công",
             "Trả về confidence dưới dạng %, có giải thích \"Độ tin cậy: X% (dựa trên Y mẫu thị trường)\"."),
            ("EXPLAIN_AI07", "Cảnh báo khi model khai báo không khớp",
             "1. Khai báo Model khác với ảnh.",
             "Định giá thành công với modelMismatch=true",
             "Banner đỏ/cam: \"Hệ thống nhận diện máy có thể là gen X, khác khai báo của bạn\"."),
            ("EXPLAIN_AI08", "Liệt kê thế hệ nhận diện được từ ảnh",
             "1. Định giá thành công.",
             "Định giá thành công",
             "Trả về detectedGeneration (vd: \"gen_14\") cùng confidence riêng cho việc nhận diện model."),
            ("EXPLAIN_AI09", "Cho phép Seller tự chỉnh giá đề xuất",
             "1. Xem kết quả định giá.\n2. Bấm \"Sử dụng giá khác\".",
             "Định giá thành công",
             "Cho phép user nhập giá khác P_final khi tạo tin đăng, có warning \"Lệch giá đề xuất X%\"."),
            ("EXPLAIN_AI10", "Hiển thị tóm tắt thị trường (market summary)",
             "1. Định giá thành công.\n2. Đọc phần marketSummary.",
             "Định giá thành công",
             "Có đoạn text giải thích: \"Có X tin đăng cùng model trong 30 ngày qua, giá trung vị Y VND\"."),
        ]
    },
    {
        "header": "PERF_SHOW Kiểm thử hiệu năng",
        "cases": [
            ("PERF_AI01", "Thời gian phản hồi với 4 ảnh",
             "1. Tải 4 ảnh ~2MB mỗi ảnh.\n2. Bấm Định giá.\n3. Đo thời gian từ click đến khi có kết quả.",
             "Truy cập trang Định giá AI",
             "Thời gian phản hồi ≤ 10 giây (P95)."),
            ("PERF_AI02", "Thời gian phản hồi với 6 ảnh (tối đa)",
             "1. Tải 6 ảnh ~3MB mỗi ảnh.\n2. Bấm Định giá.\n3. Đo thời gian.",
             "Truy cập trang Định giá AI",
             "Thời gian phản hồi ≤ 15 giây (P95). Latency YOLO inference < 500ms cho 4 ảnh."),
            ("PERF_AI03", "Cache P_market hoạt động (TTL=24h)",
             "1. Định giá iPhone 14 lần 1, đo thời gian.\n2. Định giá iPhone 14 lần 2 (cùng model), đo thời gian.",
             "Redis đang chạy",
             "Lần 2 nhanh hơn lần 1 ≥ 30% (do hit cache P_market thay vì query MongoDB)."),
            ("PERF_AI04", "10 user định giá đồng thời",
             "1. Mở 10 tab.\n2. Cùng lúc submit yêu cầu định giá.",
             "Backend, ai-service-vision sẵn sàng",
             "Tất cả request đều trả kết quả trong < 30s. Không có lỗi 500/timeout."),
            ("PERF_AI05", "Ảnh độ phân giải cao (4K, 4032x3024)",
             "1. Tải lên ảnh 4K từ iPhone.\n2. Bấm Định giá.\n3. Đo thời gian + memory.",
             "Truy cập trang Định giá AI",
             "YOLO resize về 640x640 và xử lý xong < 12 giây. Không OOM trên RTX 3050 4GB."),
            ("PERF_AI06", "Định giá khi ai-service-vision DOWN",
             "1. Tắt service ai-service-vision.\n2. Bấm Định giá.",
             "",
             "Backend trả 503 trong < 3s, frontend hiển thị \"Dịch vụ AI tạm thời không khả dụng, vui lòng thử lại sau\"."),
            ("PERF_AI07", "Định giá khi MongoDB DOWN (không có dữ liệu market)",
             "1. Tắt MongoDB.\n2. Bấm Định giá.",
             "",
             "Hệ thống fallback dùng dữ liệu cache HOẶC trả P_final = null kèm thông báo \"Không có dữ liệu thị trường\"."),
        ]
    },
]

current_row = last_row + 1
total_added = 0
for cat in categories:
    write_section_header(current_row, cat["header"])
    current_row += 1
    for case in cat["cases"]:
        write_test_case(current_row, *case)
        current_row += 1
        total_added += 1

wb.save(XLSX_PATH)
print(f"Added {total_added} test cases across {len(categories)} categories.")
print(f"New last row: {current_row - 1}")
